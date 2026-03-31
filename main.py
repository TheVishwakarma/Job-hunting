"""
JobHunter Pro — FastAPI Backend v2
Run: uvicorn main:app --reload --port 8000
"""

from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
import sqlite3, json, os, httpx, asyncio, re
from datetime import datetime
from bs4 import BeautifulSoup
import anthropic

app = FastAPI(title="JobHunter Pro API", version="2.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

CLAUDE_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
ADZUNA_APP_ID  = os.getenv("ADZUNA_APP_ID", "")
ADZUNA_API_KEY = os.getenv("ADZUNA_API_KEY", "")
DB_PATH        = "jobhunter.db"

# ── 59 companies from Sunil's tracker + additions ─────────────────────────────
SEED_COMPANIES = [
    # From job_Tracker_2024.xlsx
    {"name": "Harman", "url": "jobs.harman.com/en_US/careers", "sector": "Electronics/Tech"},
    {"name": "Cornerstone OnDemand", "url": "careers.csod.com", "sector": "HRTech"},
    {"name": "Intel", "url": "jobs.intel.com", "sector": "Semiconductor"},
    {"name": "Upstox", "url": "upstox.com/careers", "sector": "Fintech"},
    {"name": "Paytm", "url": "paytm.com/careers", "sector": "Fintech"},
    {"name": "Seagate", "url": "seagate.com/em/en/company/careers", "sector": "Hardware"},
    {"name": "Dentsu", "url": "careers.dentsu.com", "sector": "Advertising"},
    {"name": "Salesforce", "url": "salesforce.com/careers", "sector": "CRM/SaaS"},
    {"name": "Birlasoft", "url": "birlasoft.com/careers", "sector": "IT Services"},
    {"name": "Revolut", "url": "revolut.com/careers", "sector": "Fintech"},
    {"name": "BCG", "url": "careers.bcg.com", "sector": "Consulting"},
    {"name": "PwC", "url": "pwc.com/gx/en/careers.html", "sector": "Big4"},
    {"name": "EY", "url": "careers.ey.com", "sector": "Big4"},
    {"name": "Zoho", "url": "careers.zohocorp.com", "sector": "SaaS"},
    {"name": "State Street", "url": "statestreet.com/us/en/individual/careers", "sector": "Finance"},
    {"name": "IQVIA", "url": "jobs.iqvia.com", "sector": "Healthcare"},
    {"name": "Oracle", "url": "oracle.com/careers", "sector": "Enterprise Tech"},
    {"name": "Siemens", "url": "siemens.com/global/en/company/jobs", "sector": "Industrial"},
    {"name": "Walmart", "url": "careers.walmart.com", "sector": "Retail"},
    {"name": "Invesco", "url": "invesco.com/corporate/en/careers", "sector": "Finance"},
    {"name": "Ascensus", "url": "ascensus.com/about/careers", "sector": "Finance"},
    {"name": "Groww", "url": "careers.groww.in", "sector": "Fintech"},
    {"name": "Microsoft", "url": "careers.microsoft.com", "sector": "Big Tech"},
    {"name": "Wells Fargo", "url": "wellsfargojobs.com", "sector": "Banking"},
    {"name": "Nokia", "url": "nokia.com/about-us/careers", "sector": "Telecom"},
    {"name": "Pearson", "url": "pearson.com/en-us/careers.html", "sector": "EdTech"},
    {"name": "NTT Data", "url": "nttdata.com/global/en/careers", "sector": "IT Services"},
    {"name": "Target", "url": "target.com/careers", "sector": "Retail"},
    {"name": "Volvo Group", "url": "jobs.volvogroup.com", "sector": "Automotive"},
    {"name": "KPMG", "url": "kpmg.com/xx/en/home/careers.html", "sector": "Big4"},
    {"name": "Wipro", "url": "careers.wipro.com", "sector": "IT Services"},
    {"name": "Mercedes-Benz", "url": "mercedes-benz.com/en/career", "sector": "Automotive"},
    {"name": "S&P Global", "url": "careers.spglobal.com", "sector": "Finance/Data"},
    {"name": "Accenture", "url": "accenture.com/in-en/careers", "sector": "Consulting"},
    {"name": "Northern Trust", "url": "northerntrust.com/united-states/careers", "sector": "Banking"},
    {"name": "Bristol-Myers Squibb", "url": "careers.bms.com", "sector": "Pharma"},
    {"name": "Shell", "url": "careers.shell.com", "sector": "Energy"},
    {"name": "ZS Associates", "url": "zs.com/careers", "sector": "Consulting"},
    {"name": "Myntra", "url": "myntra.com/careers", "sector": "Fashion Tech"},
    {"name": "American Express", "url": "aexp.com/careers", "sector": "Fintech"},
    {"name": "Swiggy", "url": "careers.swiggy.com", "sector": "Food Tech"},
    # India startups & scaleups
    {"name": "Flipkart", "url": "careers.flipkart.com", "sector": "E-commerce"},
    {"name": "CRED", "url": "careers.cred.club", "sector": "Fintech"},
    {"name": "PhonePe", "url": "phonepe.com/careers", "sector": "Fintech"},
    {"name": "Zepto", "url": "careers.zepto.com", "sector": "Quick Commerce"},
    {"name": "Meesho", "url": "meesho.io/jobs", "sector": "E-commerce"},
    {"name": "Urban Company", "url": "urbancompany.com/careers", "sector": "Services"},
    {"name": "Zomato", "url": "zomato.com/careers", "sector": "Food Tech"},
    {"name": "Razorpay", "url": "razorpay.com/jobs", "sector": "Fintech"},
    # Analytics & consulting
    {"name": "Genpact", "url": "genpact.com/careers", "sector": "BPO/Analytics"},
    {"name": "Deloitte", "url": "careers.deloitte.com/in", "sector": "Big4"},
    {"name": "Gainwell Technologies", "url": "jobs.gainwelltechnologies.com", "sector": "Healthcare IT"},
    {"name": "Mu Sigma", "url": "mu-sigma.com/careers", "sector": "Analytics"},
    {"name": "Fractal Analytics", "url": "fractal.ai/careers", "sector": "Analytics"},
    {"name": "Tiger Analytics", "url": "tigeranalytics.com/careers", "sector": "Analytics"},
    # IT services
    {"name": "Infosys", "url": "infosys.com/careers", "sector": "IT Services"},
    {"name": "TCS", "url": "tcs.com/careers", "sector": "IT Services"},
    {"name": "HCL Technologies", "url": "hcltech.com/careers", "sector": "IT Services"},
    {"name": "Capgemini", "url": "capgemini.com/in-en/careers", "sector": "IT Services"},
    {"name": "EXL Service", "url": "exlservice.com/careers", "sector": "Analytics/BPO"},
]

# ── ATS detection ─────────────────────────────────────────────────────────────
ATS_PATTERNS = {
    "greenhouse": ["greenhouse.io", "boards.greenhouse.io"],
    "lever":      ["lever.co", "jobs.lever.co"],
    "workday":    ["myworkdayjobs.com", "wd1.myworkdayjobs", "wd3.myworkdayjobs", "wd5.myworkdayjobs"],
    "icims":      ["icims.com"],
    "taleo":      ["taleo.net"],
    "jobvite":    ["jobvite.com"],
    "smartr":     ["smartrecruiters.com"],
}

def detect_ats(url: str) -> str:
    url_lower = url.lower()
    for ats, patterns in ATS_PATTERNS.items():
        if any(p in url_lower for p in patterns):
            return ats
    return "generic"

# ── DB setup ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            url TEXT NOT NULL,
            sector TEXT DEFAULT '',
            status TEXT DEFAULT 'idle',
            jobs_found INTEGER DEFAULT 0,
            last_scan TEXT DEFAULT 'Never',
            ats_type TEXT DEFAULT 'generic',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            company TEXT NOT NULL,
            location TEXT DEFAULT 'India',
            type TEXT DEFAULT 'Full-time',
            source TEXT DEFAULT 'Career Page',
            url TEXT DEFAULT '',
            description TEXT DEFAULT '',
            score INTEGER DEFAULT 0,
            tier TEXT DEFAULT 'mid',
            skills TEXT DEFAULT '[]',
            posted TEXT DEFAULT '',
            status TEXT DEFAULT 'todo',
            notes TEXT DEFAULT '',
            applied_at TEXT DEFAULT '',
            referral TEXT DEFAULT '',
            referrer TEXT DEFAULT '',
            is_new INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS resume (
            id INTEGER PRIMARY KEY,
            filename TEXT,
            content TEXT,
            uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS scan_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER,
            company_name TEXT,
            jobs_found INTEGER DEFAULT 0,
            status TEXT,
            scanned_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)
    db.commit()
    count = db.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    if count == 0:
        for c in SEED_COMPANIES:
            db.execute(
                "INSERT INTO companies (name, url, sector, ats_type) VALUES (?,?,?,?)",
                (c["name"], c["url"], c["sector"], detect_ats(c["url"]))
            )
        db.commit()
        print(f"Seeded {len(SEED_COMPANIES)} companies")
    db.close()

init_db()

# ── Pydantic models ───────────────────────────────────────────────────────────
class Company(BaseModel):
    name: str
    url: str
    sector: Optional[str] = ""

class JobUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    applied_at: Optional[str] = None
    referral: Optional[str] = None
    referrer: Optional[str] = None

class JobManual(BaseModel):
    title: str
    company: str
    location: Optional[str] = "India"
    url: Optional[str] = ""
    description: Optional[str] = ""
    source: Optional[str] = "Manual"

class AnalyseRequest(BaseModel):
    jd_text: str
    job_url: Optional[str] = ""

class SettingUpdate(BaseModel):
    key: str
    value: str

# ── Helpers ───────────────────────────────────────────────────────────────────
def row_to_dict(row):
    return dict(row) if row else None

def rows_to_list(rows):
    result = []
    for r in rows:
        d = dict(r)
        if "skills" in d:
            try: d["skills"] = json.loads(d["skills"])
            except: d["skills"] = []
        result.append(d)
    return result

RESUME_FALLBACK = """
Sunil Kumar — Data Analyst, 1.6 years experience
Skills: SQL (MySQL), Python, Power BI, Excel, Pandas, NumPy, Scikit-learn, XGBoost, Streamlit, RAG/LLM systems, Git
Experience: Data Research Analyst at Com1UK Technologies — data validation, MySQL queries, LLM-powered Bid Assistant Q&A engine, RAG conversational agent (30-40% performance improvement)
Internship: Edunet Foundation — Power BI dashboards, restaurant analytics (100+ locations)
Project: Laptop Price Prediction ML model — 85% accuracy, XGBoost, Streamlit UI
Education: B.E. Software Engineering, IKGPTU, 8.1 GPA (2021-2024)
Certifications: Business Analytics & Text Mining (NPTEL), Python for Data Science (IBM), SQL Advanced + Python Basic (Udemy)
Target roles: Junior Data Analyst, BI Analyst, MIS Analyst, Business Analyst, Data Research Analyst
Location: India (Bangalore preferred)
"""

def get_resume_text() -> str:
    db = get_db()
    row = db.execute("SELECT content FROM resume LIMIT 1").fetchone()
    db.close()
    return (row["content"] if row else "").strip() or RESUME_FALLBACK

# ── Resume ────────────────────────────────────────────────────────────────────
@app.post("/api/resume/upload")
async def upload_resume(file: UploadFile = File(...)):
    content = await file.read()
    try:
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(content))
        text = "\n".join(p.extract_text() or "" for p in reader.pages)
    except:
        text = content.decode("utf-8", errors="ignore")
    db = get_db()
    db.execute("DELETE FROM resume")
    db.execute("INSERT INTO resume (filename, content) VALUES (?,?)", (file.filename, text))
    db.commit()
    db.close()
    return {"status": "ok", "filename": file.filename, "chars": len(text)}

@app.get("/api/resume")
def get_resume():
    db = get_db()
    row = db.execute("SELECT id, filename, uploaded_at FROM resume ORDER BY id DESC LIMIT 1").fetchone()
    db.close()
    return row_to_dict(row) or {}

# ── Import from Excel tracker ─────────────────────────────────────────────────
@app.post("/api/import/tracker")
def import_tracker():
    xlsx_path = "job_Tracker_2024.xlsx"
    if not os.path.exists(xlsx_path):
        raise HTTPException(404, f"{xlsx_path} not found. Place it in the backend folder.")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb['Company tracker']
    except Exception as e:
        raise HTTPException(500, f"Could not read Excel: {e}")

    STATUS_MAP = {
        "Mail Received": "applied", "Awaited": "watching",
        "Rejected": "rejected", "EXPIRED": "rejected",
        "NA": "todo", "can't refer": "todo", "None": "todo",
    }
    db = get_db()
    imported = skipped = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3: continue
        if not row[0] or str(row[0]).strip() in ['', 'Company Name']: continue
        raw_name  = str(row[0]).strip()
        job_type  = str(row[1]).strip() if row[1] else "Analyst"
        referral  = str(row[3]).strip() if row[3] else ""
        referrer  = str(row[5]).strip() if row[5] else ""
        applied   = str(row[7]).strip() if row[7] else ""
        applied_dt= str(row[8])[:10] if row[8] and str(row[8]) != "None" else ""
        raw_status= str(row[10]).strip() if row[10] else "Awaited"
        job_link  = str(row[13]).strip() if row[13] and str(row[13]) != "None" else ""
        company   = raw_name.split(',')[0].strip()
        status    = STATUS_MAP.get(raw_status, "todo")
        if applied == "Yes" and status == "todo":
            status = "applied"
        existing = db.execute(
            "SELECT id FROM jobs WHERE title=? AND company=? AND source='Tracker'",
            (job_type, company)
        ).fetchone()
        if existing:
            skipped += 1
            continue
        db.execute("""
            INSERT INTO jobs (title, company, location, source, url, score, tier,
                              skills, posted, status, referral, referrer, applied_at, is_new)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            job_type, company, "India", "Tracker",
            job_link, 0, "mid", "[]",
            applied_dt or "2024", status,
            referral, referrer, applied_dt, 0
        ))
        imported += 1
    db.commit()
    db.close()
    return {"status": "ok", "imported": imported, "skipped": skipped,
            "message": f"Imported {imported} entries from your tracker. Duplicates skipped: {skipped}."}

# ── Companies ─────────────────────────────────────────────────────────────────
@app.get("/api/companies")
def list_companies(sector: str = None, status: str = None):
    db = get_db()
    q = "SELECT * FROM companies WHERE 1=1"
    p = []
    if sector: q += " AND sector=?"; p.append(sector)
    if status: q += " AND status=?"; p.append(status)
    q += " ORDER BY name"
    rows = db.execute(q, p).fetchall()
    db.close()
    return rows_to_list(rows)

@app.get("/api/companies/sectors")
def list_sectors():
    db = get_db()
    rows = db.execute("SELECT DISTINCT sector FROM companies ORDER BY sector").fetchall()
    db.close()
    return [r[0] for r in rows]

@app.post("/api/companies")
def add_company(c: Company):
    db = get_db()
    cur = db.execute(
        "INSERT INTO companies (name, url, sector, ats_type) VALUES (?,?,?,?)",
        (c.name, c.url, c.sector, detect_ats(c.url))
    )
    db.commit()
    row = db.execute("SELECT * FROM companies WHERE id=?", (cur.lastrowid,)).fetchone()
    db.close()
    return row_to_dict(row)

@app.delete("/api/companies/{company_id}")
def delete_company(company_id: int):
    db = get_db()
    db.execute("DELETE FROM companies WHERE id=?", (company_id,))
    db.commit()
    db.close()
    return {"status": "deleted"}

# ── Smart scraper ─────────────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

TARGET_KEYWORDS = [
    "data analyst", "business analyst", "bi analyst", "analytics",
    "mis analyst", "reporting analyst", "data research", "intelligence analyst",
    "junior analyst", "associate analyst", "python analyst", "sql analyst",
    "data science", "junior data", "analytics engineer",
]

def is_relevant(title: str) -> bool:
    t = title.lower()
    return any(k in t for k in TARGET_KEYWORDS)

async def scrape_greenhouse(url: str, name: str) -> List[dict]:
    slug = re.search(r'greenhouse\.io/([^/?#\s]+)', url)
    board = slug.group(1).split('/')[0] if slug else name.lower().replace(' ','-')
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.get(f"https://boards-api.greenhouse.io/v1/boards/{board}/jobs?content=true")
        return [{
            "title": j.get("title",""), "company": name,
            "url": j.get("absolute_url",""), "source": "Career Page",
            "location": j.get("location",{}).get("name","India"),
            "description": BeautifulSoup(j.get("content",""),"html.parser").get_text()[:600],
        } for j in r.json().get("jobs",[])]
    except:
        return await scrape_generic(url, name)

async def scrape_lever(url: str, name: str) -> List[dict]:
    slug = re.search(r'lever\.co/([^/?#\s]+)', url)
    board = slug.group(1).split('/')[0] if slug else name.lower().replace(' ','-')
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.get(f"https://api.lever.co/v0/postings/{board}?mode=json")
        return [{
            "title": j.get("text",""), "company": name,
            "url": j.get("hostedUrl",""), "source": "Career Page",
            "location": j.get("categories",{}).get("location","India"),
            "description": j.get("descriptionPlain","")[:600],
        } for j in r.json()]
    except:
        return await scrape_generic(url, name)

async def scrape_generic(url: str, name: str) -> List[dict]:
    full = f"https://{url}" if not url.startswith("http") else url
    jobs, seen = [], set()
    try:
        async with httpx.AsyncClient(timeout=18, follow_redirects=True, headers=HEADERS) as c:
            resp = await c.get(full)
        soup = BeautifulSoup(resp.text, "html.parser")
        for t in soup(["script","style","nav","footer","header"]): t.decompose()
        for sel in [
            "h2.job-title","h3.job-title",".job-listing h2",".job-listing h3",
            ".position-title",".opening-title",".job-card h3","[data-job-title]",
            "a[href*='/jobs/']","a[href*='/careers/']","a[href*='/openings/']",
        ]:
            for tag in soup.select(sel)[:30]:
                text = tag.get_text(strip=True)
                if not text or len(text)<5 or len(text)>120 or text in seen: continue
                seen.add(text)
                href = tag.get("href","") if tag.name=="a" else ""
                if href and not href.startswith("http"):
                    href = "/".join(full.split("/")[:3]) + "/" + href.lstrip("/")
                jobs.append({"title":text,"company":name,"url":href or full,
                             "source":"Career Page","location":"India","description":""})
        # Fallback — look for any anchor that mentions a relevant job
        if not jobs:
            for a in soup.find_all("a", href=True)[:100]:
                text = a.get_text(strip=True)
                if 5<len(text)<100 and is_relevant(text) and text not in seen:
                    seen.add(text)
                    href = a["href"]
                    if not href.startswith("http"):
                        href = "/".join(full.split("/")[:3])+"/"+href.lstrip("/")
                    jobs.append({"title":text,"company":name,"url":href,
                                 "source":"Career Page","location":"India","description":""})
    except Exception as e:
        print(f"  scrape_generic error [{name}]: {e}")
    return jobs

async def scrape_company(company: dict) -> List[dict]:
    ats = company.get("ats_type","generic")
    url, name = company["url"], company["name"]
    if ats == "greenhouse": return await scrape_greenhouse(url, name)
    if ats == "lever":      return await scrape_lever(url, name)
    return await scrape_generic(url, name)

# ── Claude Haiku scorer ───────────────────────────────────────────────────────
async def score_job(title: str, description: str, resume_text: str) -> dict:
    kws = ["sql","python","power bi","data analyst","pandas","ml","excel",
           "analytics","bi","analyst","mis","reporting","rag","llm","streamlit",
           "scikit","xgboost","numpy","business analyst","kpi","dashboard"]
    combo = (title+" "+description).lower()
    hits = sum(1 for k in kws if k in combo)
    score = min(45+hits*6, 94)
    tier  = "high" if score>=80 else "mid" if score>=65 else "stretch"
    if not CLAUDE_API_KEY:
        return {"score":score,"tier":tier,"matched_skills":[],"missing_skills":[]}
    try:
        client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=280,
            messages=[{"role":"user","content":
                f"""Score this job vs resume. Return ONLY JSON, no markdown.
RESUME: {resume_text[:1500]}
JOB: {title} — {description[:600]}
JSON: {{"score":<0-100>,"tier":"<high|mid|stretch>","matched_skills":["..."],"missing_skills":["..."]}}
high=80+,mid=65-79,stretch<65"""}]
        )
        raw = re.sub(r"```json|```","",msg.content[0].text).strip()
        return json.loads(raw)
    except Exception as e:
        print(f"  Claude score error: {e}")
        return {"score":score,"tier":tier,"matched_skills":[],"missing_skills":[]}

# ── Scan ──────────────────────────────────────────────────────────────────────
@app.post("/api/scan")
async def run_scan(background_tasks: BackgroundTasks, company_id: Optional[int] = None):
    background_tasks.add_task(scan_companies, company_id)
    return {"status":"scan_started","target": f"company #{company_id}" if company_id else "all companies"}

async def scan_companies(company_id: Optional[int] = None):
    db = get_db()
    companies = rows_to_list(
        db.execute("SELECT * FROM companies WHERE id=?", (company_id,)).fetchall()
        if company_id else
        db.execute("SELECT * FROM companies").fetchall()
    )
    resume_text = get_resume_text()
    db.close()

    for company in companies:
        db = get_db()
        try:
            db.execute("UPDATE companies SET status='active' WHERE id=?", (company["id"],))
            db.commit()
            raw = await scrape_company(company)
            relevant = [j for j in raw if is_relevant(j["title"])] or raw[:5]
            new_count = 0
            for rj in relevant[:12]:
                if db.execute("SELECT id FROM jobs WHERE title=? AND company=?",
                              (rj["title"],rj["company"])).fetchone():
                    continue
                res = await score_job(rj["title"],rj.get("description",""),resume_text)
                db.execute("""INSERT INTO jobs (title,company,location,source,url,description,
                    score,tier,skills,posted,status,is_new) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (rj["title"],rj["company"],rj.get("location","India"),"Career Page",
                     rj.get("url",""),rj.get("description","")[:500],
                     res["score"],res.get("tier","mid"),
                     json.dumps(res.get("matched_skills",[])),
                     datetime.now().strftime("%b %d"),"todo",1))
                new_count += 1
            db.execute("UPDATE companies SET status='active',jobs_found=?,last_scan=? WHERE id=?",
                (new_count, datetime.now().strftime("%d %b, %H:%M"), company["id"]))
            db.execute("INSERT INTO scan_log (company_id,company_name,jobs_found,status) VALUES (?,?,?,?)",
                (company["id"],company["name"],new_count,"ok"))
            db.commit()
            print(f"  {company['name']}: {new_count} new jobs")
        except Exception as e:
            print(f"  Error [{company['name']}]: {e}")
            db.execute("UPDATE companies SET status='error',last_scan='Failed' WHERE id=?", (company["id"],))
            db.execute("INSERT INTO scan_log (company_id,company_name,jobs_found,status) VALUES (?,?,?,?)",
                (company["id"],company["name"],0,f"error:{str(e)[:80]}"))
            db.commit()
        finally:
            db.close()
        await asyncio.sleep(1.5)

# ── Adzuna ────────────────────────────────────────────────────────────────────
@app.post("/api/adzuna/fetch")
async def fetch_adzuna(role:str="data analyst", location:str="india", pages:int=3):
    if not ADZUNA_APP_ID:
        raise HTTPException(400,"Adzuna API keys not set in Settings")
    resume_text = get_resume_text()
    db = get_db()
    added = 0
    async with httpx.AsyncClient(timeout=20) as client:
        for page in range(1, pages+1):
            try:
                r = await client.get(f"https://api.adzuna.com/v1/api/jobs/in/search/{page}",
                    params={"app_id":ADZUNA_APP_ID,"app_key":ADZUNA_API_KEY,
                            "results_per_page":20,"what":role,"where":location,"sort_by":"date"})
                for j in r.json().get("results",[]):
                    t,co = j.get("title",""),j.get("company",{}).get("display_name","")
                    if db.execute("SELECT id FROM jobs WHERE title=? AND company=?",(t,co)).fetchone(): continue
                    res = await score_job(t,j.get("description","")[:600],resume_text)
                    db.execute("""INSERT INTO jobs (title,company,location,source,url,description,
                        score,tier,skills,posted,status,is_new) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (t,co,j.get("location",{}).get("display_name","India"),"Adzuna",
                         j.get("redirect_url",""),j.get("description","")[:400],
                         res["score"],res.get("tier","mid"),
                         json.dumps(res.get("matched_skills",[])),
                         datetime.now().strftime("%b %d"),"todo",1))
                    added += 1
            except Exception as e:
                print(f"Adzuna p{page}: {e}")
    db.commit(); db.close()
    return {"status":"ok","jobs_added":added}

# ── Manual job add ─────────────────────────────────────────────────────────────
@app.post("/api/jobs/manual")
async def add_job_manual(job: JobManual):
    desc = job.description
    if job.url and not desc:
        try:
            async with httpx.AsyncClient(timeout=12,follow_redirects=True,headers=HEADERS) as c:
                r = await c.get(job.url)
            soup = BeautifulSoup(r.text,"html.parser")
            for t in soup(["script","style","nav","footer"]): t.decompose()
            desc = soup.get_text(" ",strip=True)[:800]
        except: pass
    resume_text = get_resume_text()
    res = await score_job(job.title, desc, resume_text)
    db = get_db()
    cur = db.execute("""INSERT INTO jobs (title,company,location,source,url,description,
        score,tier,skills,posted,status,is_new) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (job.title,job.company,job.location or "India",job.source or "Manual",
         job.url or "",desc[:400],res["score"],res.get("tier","mid"),
         json.dumps(res.get("matched_skills",[])),
         datetime.now().strftime("%b %d"),"todo",1))
    db.commit()
    row = db.execute("SELECT * FROM jobs WHERE id=?", (cur.lastrowid,)).fetchone()
    db.close()
    return row_to_dict(row)

# ── Jobs CRUD ─────────────────────────────────────────────────────────────────
@app.get("/api/jobs")
def list_jobs(tier:str=None,status:str=None,source:str=None,q:str=None,
              min_score:int=None,limit:int=300):
    db = get_db()
    query,params = "SELECT * FROM jobs WHERE 1=1", []
    if tier:      query+=" AND tier=?";       params.append(tier)
    if status:    query+=" AND status=?";     params.append(status)
    if source:    query+=" AND source=?";     params.append(source)
    if min_score: query+=" AND score>=?";     params.append(min_score)
    if q:
        query+=" AND (title LIKE ? OR company LIKE ? OR skills LIKE ?)"
        params+=[f"%{q}%",f"%{q}%",f"%{q}%"]
    query+=f" ORDER BY score DESC, created_at DESC LIMIT {limit}"
    rows = db.execute(query,params).fetchall()
    db.close()
    return rows_to_list(rows)

@app.patch("/api/jobs/{job_id}")
def update_job(job_id:int, update:JobUpdate):
    db = get_db()
    fields,vals = [],[]
    for f,v in update.dict(exclude_none=True).items():
        fields.append(f"{f}=?"); vals.append(v)
    if fields:
        db.execute(f"UPDATE jobs SET {','.join(fields)} WHERE id=?", vals+[job_id])
        db.commit()
    row = db.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    db.close()
    return row_to_dict(row)

@app.delete("/api/jobs/{job_id}")
def delete_job(job_id:int):
    db = get_db()
    db.execute("DELETE FROM jobs WHERE id=?", (job_id,))
    db.commit(); db.close()
    return {"status":"deleted"}

@app.post("/api/jobs/mark-seen")
def mark_all_seen():
    db = get_db()
    db.execute("UPDATE jobs SET is_new=0")
    db.commit(); db.close()
    return {"status":"ok"}

# ── Gap analysis ──────────────────────────────────────────────────────────────
@app.post("/api/analyse")
async def analyse_fit(req: AnalyseRequest):
    jd = req.jd_text
    if req.job_url and not jd:
        try:
            async with httpx.AsyncClient(timeout=12,follow_redirects=True,headers=HEADERS) as c:
                r = await c.get(req.job_url)
            soup = BeautifulSoup(r.text,"html.parser")
            for t in soup(["script","style","nav","footer"]): t.decompose()
            jd = soup.get_text(" ",strip=True)[:2500]
        except: pass
    if not jd.strip():
        raise HTTPException(400,"No JD text provided")
    resume_text = get_resume_text()
    if not CLAUDE_API_KEY:
        return {"score":72,"match_summary":"Set Claude API key in Settings for AI analysis.",
                "matched_skills":["SQL","Python","Power BI"],"missing_skills":["Tableau","AWS"],
                "resume_suggestions":["Add your Claude API key in Settings"]}
    try:
        client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=900,
            messages=[{"role":"user","content":
                f"""Expert recruiter doing resume-JD analysis. Return ONLY valid JSON, no markdown.
RESUME: {resume_text[:2500]}
JOB DESCRIPTION: {jd[:2000]}
Return: {{"score":<0-100>,"match_summary":"<2-3 sentences>","matched_skills":["..."],"missing_skills":["..."],"resume_suggestions":["tip1","tip2","tip3"]}}"""}]
        )
        raw = re.sub(r"```json|```","",msg.content[0].text).strip()
        return json.loads(raw)
    except Exception as e:
        raise HTTPException(500,f"Analysis failed: {e}")

# ── Settings ──────────────────────────────────────────────────────────────────
@app.get("/api/settings")
def get_settings():
    db = get_db()
    rows = db.execute("SELECT * FROM settings").fetchall()
    db.close()
    return {r["key"]:r["value"] for r in rows}

@app.post("/api/settings")
def save_setting(s: SettingUpdate):
    global CLAUDE_API_KEY, ADZUNA_APP_ID, ADZUNA_API_KEY
    db = get_db()
    db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (s.key,s.value))
    db.commit(); db.close()
    if s.key=="claude_api_key":  CLAUDE_API_KEY=s.value
    if s.key=="adzuna_app_id":   ADZUNA_APP_ID=s.value
    if s.key=="adzuna_api_key":  ADZUNA_API_KEY=s.value
    return {"status":"saved"}

# ── Stats & scan log ──────────────────────────────────────────────────────────
@app.get("/api/stats")
def get_stats():
    db = get_db()
    total      = db.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
    applied    = db.execute("SELECT COUNT(*) FROM jobs WHERE status='applied'").fetchone()[0]
    interviews = db.execute("SELECT COUNT(*) FROM jobs WHERE status='interview'").fetchone()[0]
    rejected   = db.execute("SELECT COUNT(*) FROM jobs WHERE status='rejected'").fetchone()[0]
    watching   = db.execute("SELECT COUNT(*) FROM jobs WHERE status='watching'").fetchone()[0]
    new_jobs   = db.execute("SELECT COUNT(*) FROM jobs WHERE is_new=1").fetchone()[0]
    companies  = db.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    avg_row    = db.execute("SELECT AVG(score) FROM jobs WHERE score>0").fetchone()
    avg_score  = round(avg_row[0] or 0)
    high_prob  = db.execute("SELECT COUNT(*) FROM jobs WHERE tier='high'").fetchone()[0]
    trend      = rows_to_list(db.execute(
        "SELECT date(scanned_at) as day, SUM(jobs_found) as jobs FROM scan_log GROUP BY day ORDER BY day DESC LIMIT 7"
    ).fetchall())
    db.close()
    return {"total_jobs":total,"applied":applied,"interviews":interviews,"rejected":rejected,
            "watching":watching,"new_jobs":new_jobs,"companies":companies,"avg_score":avg_score,
            "high_prob":high_prob,"response_rate":round((interviews/applied*100) if applied else 0),
            "weekly_trend":trend[::-1]}

@app.get("/api/scan-log")
def get_scan_log(limit:int=50):
    db = get_db()
    rows = db.execute("SELECT * FROM scan_log ORDER BY scanned_at DESC LIMIT ?", (limit,)).fetchall()
    db.close()
    return rows_to_list(rows)

@app.get("/")
def root():
    return {"app":"JobHunter Pro API v2","companies":len(SEED_COMPANIES),"docs":"/docs"}
